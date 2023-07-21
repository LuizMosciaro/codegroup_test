import logging
import os
import smtplib
from dataclasses import dataclass
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from time import sleep
from typing import Optional

import requests
from dotenv import load_dotenv
from fake_useragent import UserAgent
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.firefox import GeckoDriverManager

load_dotenv()

@dataclass
class Pessoa:
    nome: str
    cidade: str
    estado: str
    descricao: str
    cargo: str
    empresa: str
    clima: str
    
class Crawler:
    def __init__(self,url:str,login:str,password:str,file:str,headless:Optional[bool]=None) -> None:
        self.url = url
        self.login = login
        self.password = password
        self.headless = headless
        self.file = file

    def run(self,):
        '''É a funcao principal, método responsavel por organizar e chamar os demais métodos e iniciar a automação'''

        print('Inicializando busca de informações de usuários')
        
        #Chamada para instanciar o logger e o driver
        logger = self.get_logger()
        driver = self.get_driver(self.headless,False)
        
        #Tentara 5 vezes chamar o login ou enquanto o retorno de success nao for verdadeiro
        success = False
        tries = 0
        while not success and tries <=5:
            success = self.login_site(driver, logger, self.url, self.login, self.password)
            tries += 1

        if self.get_users_information(driver,logger,self.file):
            self.send_email(os.getenv('RECIPIENT'),'Pesquisa de usuarios','Ola, segue os arquivos em anexo',self.file)

    def get_logger(self,):
        '''Logger para registro de erros, bugs e falhas'''

        logging.basicConfig(level=logging.WARNING,filename='file.log')
        log = logging.getLogger('Log')
        return log
            
    def get_driver(self, headless: bool = False, maximize: bool = True):
        '''
        Configurações do driver e inicialização da instância.

        Parâmetros:
            headless (bool): Se True, o navegador será iniciado em modo invisível (headless mode).
            maximize (bool): Se True, a janela do navegador será maximizada ao iniciar.

        Retorno:
            webdriver.Firefox: Instância do driver do navegador Firefox configurada.

        Notas:
            Este método configura o driver do navegador Firefox com as opções e preferências desejadas.
            Se a opção headless for True, o navegador será iniciado em modo invisível.
            Caso maximize seja True, a janela do navegador será maximizada ao iniciar.
            Além disso, este método configura um proxy para o navegador usando as informações fornecidas.
            Certifique-se de fornecer as informações corretas do proxy (username, password e GEONODE_DNS).
            O Geckodriver é configurado através do Manager, o que evita a necessidade de fazer o download manual do driver.

        Exemplo:
            driver = get_driver(headless=True, maximize=False)
            driver.get('https://www.example.com')
        '''
        firefox_options = Options()
        if headless:
            firefox_options.add_argument('-headless')  # Modo invisível

        # Obter um user agent aleatório
        ua = UserAgent()
        user_agent = ua.random

        #Configurações para ignorar certs, ajuste de gpu e nivel de logs
        firefox_options.add_argument('--ignore-certificate-errors')
        firefox_options.add_argument('--log-level=3')
        firefox_options.add_argument('--disable-gpu')
        firefox_options.set_preference("browser.download.folderList", 2)

        pasta_download_relativa = str(os.path.join(Path.home(), "Downloads"))
        firefox_options.set_preference("browser.download.dir", pasta_download_relativa)
        firefox_options.set_preference("browser.download.useDownloadDir", True)
        firefox_options.set_preference("browser.helperApps.neverAsk.saveToDisk", "application/pdf")
        firefox_options.set_preference("general.useragent.override", user_agent)
        
        # Configuração para o Geckodriver atraves do Manager, nao sendo necessario download do driver
        service = Service(GeckoDriverManager().install())
        driver = webdriver.Firefox(service=service, options=firefox_options)
        if maximize:
            driver.maximize_window()

        print('Configurações do webdriver realizadas')
        return driver

    def read_file(self,logger:logging.Logger,file:str):
        '''
        Realiza a leitura do arquivo fornecido para coletar informações dos usuários e devolve uma lista de instâncias da classe Pessoa.

        Parâmetros:
            logger (logging.Logger): Instância do logger para registrar informações e erros.
            file (str): Caminho do arquivo que contém as informações dos usuários.

        Retorno:
            List[Pessoa]: Lista de instâncias da classe Pessoa, contendo as informações dos usuários lidas a partir do arquivo.

        Notas:
            Este método lê o arquivo fornecido e coleta as informações dos usuários.
            O arquivo deve conter as informações dos usuários nas colunas, com cada linha representando um usuário.
            O método cria uma instância da classe Pessoa para cada linha e retorna uma lista de usuários.
            Caso ocorra um erro na leitura do arquivo, uma mensagem de erro será registrada no logger e a função retornará uma lista vazia.
        
        Exemplo:
            users_list = read_file(my_logger, 'caminho/do/arquivo/usuarios.xlsx')
            for user in users_list:
                print(user.nome, user.cidade, user.estado)
        '''
        try:
            #Le o arquivo
            workbook = load_workbook(file)
            ws = workbook.active

            #Após iterar por todas as colunas e receber os valores, desacopla a row em uma classe Pessoa e retorna uma lista
            # de pessoas
            users = []
            for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
                users.append(Pessoa(*row))

            return users
        
        except Exception as Err:
            print(Err)
            logger.warning(f'Erro ao realizar leitura do arquivo: {file}\nErro: {Err}')

    def get_temperature(self,city:str,api_key:str):
        '''
        Realiza uma chamada à API Weatherapi para obter a temperatura atual de uma cidade.

        Parâmetros:
            city (str): Nome da cidade desejada para obter a temperatura.
            api_key (str): Chave de acesso à API Weatherapi.

        Retorno:
            str: Temperatura atual em graus Celsius na cidade especificada, ou uma mensagem de erro em caso de falha.

        Notas:
            Este método utiliza a API Weatherapi para obter a temperatura atual de uma cidade.
            Certifique-se de fornecer o nome da cidade usando a nomenclatura padrão com a ISO internacional (exemplo: São Paulo, BR).
            A chave da API Weatherapi é necessária para acessar os dados climáticos.
            Em caso de sucesso, retorna a temperatura atual em graus Celsius.
            Em caso de falha na chamada da API, retorna uma mensagem de erro com o status do código de resposta.

        Exemplo:
            temperatura = get_temperature('São Paulo', 'minha_api_key')
            if temperatura:
                print(f'Temperatura em São Paulo: {temperatura}°C')
            else:
                print('Erro ao obter dados de clima.')
        '''
        url = f"http://api.weatherapi.com/v1/current.json?key={api_key}&q={city},BR"

        response = requests.get(url)

        if response.status_code == 200:
            temperatura = response.json()['current']['temp_c']
            return str(temperatura)
        else:
           return f"Erro ao obter dados de clima: {response.status_code}"
    
    def login_site(self, driver: webdriver.Chrome, logger: logging.Logger, url: str, login: str, password: str) -> bool:
        '''
        Realiza o login do usuário no site usando as credenciais fornecidas.

        Parâmetros:
            driver (webdriver.Chrome): Instância do driver do navegador Chrome, usado para acessar o site.
            logger (logging.Logger): Instância do logger para registrar informações e erros.
            url (str): URL do site de login.
            login (str): Nome de usuário ou e-mail do usuário.
            password (str): Senha do usuário.

        Retorno:
            bool: Retorna True se o login for bem-sucedido, caso contrário, False.

        Notas:
            Este método tentará realizar o login no site com as credenciais fornecidas.
            Caso a página de login mude, ele tentará novamente clicando no botão de login alternativo.
            O login é considerado bem-sucedido se o elemento 'search-global-typeahead__input' estiver presente na página após o login.
            O método tentará o login até 5 vezes, aguardando 5 segundos entre as tentativas.
        
        Exemplo:
            if login_site(my_driver, my_logger, 'https://www.example.com/login', 'meu_usuario', 'minha_senha'):
                print('Login bem-sucedido.')
        '''        
        print(f'Realizando Login: {url}')
        success = False
        tries = 0
        while not success and tries <= 5:
            try:
                driver.get(url)
                sleep(5)
                try:
                    # Credenciais
                    driver.find_element(By.ID, 'session_key').send_keys(login)
                    driver.find_element(By.ID, 'session_password').send_keys(password)
                    driver.find_element(By.XPATH, '//button[normalize-space()="Entrar"]').click()
                    success = True
                except:
                    # Caso o link mude, retorna ao clicar no botao abaixo e tentar novo login
                    driver.find_element(By.XPATH, '//button[contains(@class,"authwall-join-form__form-toggle--bottom form-toggle")]').click()
                    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, 'session_key')))
                    driver.find_element(By.ID, 'session_key').send_keys(login)
                    driver.find_element(By.ID, 'session_password').send_keys(password)
                    driver.find_element(By.XPATH, '//button[normalize-space()="Entrar"]').click()
                    success = True
                    
                #Aguardar a barra de search para input de informacoes
                WebDriverWait(driver, 120).until(
                    EC.presence_of_element_located((By.CLASS_NAME, 'search-global-typeahead__input')))
                print('Usuário conectado')
        
            except NoSuchElementException:
                logger.warning(f'Elemento não encontrado. Tentativa {tries} falhou.')
                success = False

            tries += 1
        return success

    def get_users_information(self,driver:webdriver.Chrome,logger:logging.Logger,file:str) -> None:
        '''
        Realiza o tratamento de coleta de informações principais dos usuários e as salva em um arquivo Excel.

        Parâmetros:
            driver (webdriver.Chrome): Instância do driver do navegador Chrome, usado para navegar no LinkedIn.
            logger (logging.Logger): Instância do logger para registrar informações e erros.
            file (str): Caminho do arquivo Excel onde as informações serão salvas.

        Retorno:
            None

        Notas:
            Este método coleta informações de usuários a partir de um arquivo, busca os dados dos usuários no LinkedIn e preenche
            informações como descrição, empresa, cargo e clima na cidade do usuário. As informações são salvas em um arquivo Excel.
            Certifique-se de ter a variável de ambiente 'WEATHER_API_KEY' definida, pois ela é usada para acessar dados climáticos.

        Exemplo:
            get_users_information(my_driver, my_logger, 'caminho/do/arquivo/dados_usuarios.xlsx')
        '''
        try:
            users = self.read_file(logger,file)
            users_filled_list = []
            success = True

            #Para cada usuário realizará uma pesquisa individual
            for user in users:
                input_box = driver.find_element(By.CLASS_NAME,'search-global-typeahead__input')
                input_box.send_keys(user.nome)
                input_box.send_keys(Keys.ENTER)

                #Confirmar aba 'Pessoas'
                WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH,'//button[normalize-space()="Pessoas"]')))
                driver.find_element(By.XPATH,'//button[normalize-space()="Pessoas"]').click()

                #Acessando o perfil do usuario
                print('Acessando perfil: ',user.nome)
                WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.CLASS_NAME,'entity-result__item')))
                
                #Para o caso do Tiago, como ele é uma página é necessário encontrar ele no terceiro elemento após confimar no botão 'Pessoas' do linkedin
                if user.nome == 'Tiago Pamplona':
                    elements = driver.find_elements(By.CLASS_NAME,'entity-result__item')
                    if len(elements) >= 3:
                        third_element = elements[2]
                        third_element.click()
                else:
                    driver.find_element(By.CLASS_NAME,'entity-result__item').click()

                #Aguarda a pagina carregar
                WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH,"//h2/span[contains(text(),'Sobre')][1]")))

                #Descrição
                descricao_text = driver.find_element(By.XPATH,'//div[contains(@class,"pv-shared-text-with-see-more full-width t-14 t-normal t-black display-flex align-items-center")]').text
                user.descricao = descricao_text.split('\n')[0]
                #Empresa
                empresa_text = driver.find_element(By.XPATH,'//span[contains(@class,"pv-text-details__right-panel-item-text")]').text
                user.empresa = empresa_text
                #Cargo
                cargo_text = driver.find_element(By.XPATH,'//div[contains(@class,"text-body-medium break-words")]').text
                user.cargo = cargo_text

                #Clima na cidade do usuario
                temperatura = self.get_temperature(user.cidade,os.getenv('WEATHER_API_KEY'))
                user.clima = temperatura + 'C'

                #Anexa em uma lista os objetos
                users_filled_list.append(user)
            
            driver.close()

            #Abre novamente o excel, gravando as novas informações
            book = load_workbook(file)
            worksheet = book.active
            row = 2
            
            #Para cada objeto na lista, em sua linha e coluna colocará a informação correspondente
            for user in users_filled_list:
                worksheet.cell(row=row, column=1).value = user.nome
                worksheet.cell(row=row, column=2).value = user.cidade
                worksheet.cell(row=row, column=3).value = user.estado
                worksheet.cell(row=row, column=4).value = user.descricao
                worksheet.cell(row=row, column=5).value = user.cargo
                worksheet.cell(row=row, column=6).value = user.empresa
                worksheet.cell(row=row, column=7).value = user.clima
                row += 1
            
            book.save(file)
            print(f'Dados salvos em: {file}')
            success = True

        except Exception as Err:
            print(Err)
            logger.warning(f'Erro ao pesquisar dados de usuario\nErro:{Err}')
            success = False
        
        return success
    
    def send_email(self, recipient: str, subject: str, text: str, file: str) -> None:
        '''
        Envia um e-mail com um anexo para o destinatário especificado.

        Parâmetros:
            recipient (str): Endereço de e-mail do destinatário.
            subject (str): Assunto do e-mail.
            text (str): Corpo do e-mail (texto plano).
            file (str): Caminho para o arquivo a ser anexado no e-mail.

        Retorno:
            None

        Notas:
            Certifique-se de definir as variáveis de ambiente EMAIL e EMAIL_PWD
            contendo o endereço de e-mail do remetente e a senha, respectivamente.

        Exemplo:
            send_email('destinatario@example.com', 'Assunto do E-mail', 'Corpo do e-mail em texto simples.', 'caminho/do/arquivo/anexo.txt')
        '''
        remetente = os.getenv('EMAIL')
        senha_remetente = os.getenv('EMAIL_PWD')

        # Criação da mensagem de e-mail
        msg = MIMEMultipart()
        msg['From'] = remetente
        msg['To'] = recipient
        msg['Subject'] = subject

        # Adiciona o corpo do e-mail (texto plano)
        msg.attach(MIMEText(text, 'plain'))

        # Abre o arquivo anexo em modo binário
        with open(file, 'rb') as anexo:
            attachment = MIMEBase('application', 'octet-stream')
            attachment.set_payload(anexo.read())

            # Codifica o anexo em base64
            encoders.encode_base64(attachment)
            filename = os.path.basename(file)
            attachment.add_header('Content-Disposition', f'attachment; filename={filename}')
            msg.attach(attachment)

        # Configuração do servidor SMTP do Outlook
        servidor_smtp = 'smtp.office365.com'
        porta_smtp = 587

        # Inicializa o serviço e envia o e-mail
        with smtplib.SMTP(servidor_smtp, porta_smtp) as servidor:
            servidor.starttls()
            servidor.login(remetente, senha_remetente)
            servidor.send_message(msg)
            print('E-mail enviado com sucesso')



if __name__=='__main__':
    rpa = Crawler('https://www.linkedin.com/?original_referer=',os.getenv('EMAIL'),os.getenv('PASSWORD'),'Colaboradores.xlsx',False)
    rpa.run()
